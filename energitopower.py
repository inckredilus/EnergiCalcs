import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

def calculate_energy_and_power(file_path):
    input_sheet = "InputData"
    output_sheet = "ProcessedData"

    # Läs in data
    df = pd.read_excel(file_path, sheet_name=input_sheet)

    # Säkerställ att datum är rätt typ
    df["Start"] = pd.to_datetime(df["Start"])
    df["End"] = pd.to_datetime(df["End"])

    # Beräkna Duration och Power
    df["Duration"] = (df["End"] - df["Start"]).dt.total_seconds() / 3600  # Timmar
    df["Power"] = df["Consumption"] / df["Duration"]

    # Hantera fall där perioden är kortare än en timme
    same_hour = df["Start"].dt.hour == df["End"].dt.hour
    df["Start Hour"] = df["Start"].dt.ceil("h").where(~same_hour, df["Start"])
    df["End Hour"] = df["End"].dt.floor("h").where(~same_hour, df["End"])

    # Beräkna fulla timmar och minuter före/efter
    df["Full Hours"] = ((df["End Hour"] - df["Start Hour"]).dt.total_seconds() / 3600).clip(lower=0)
    df["Before Minutes"] = ((df["Start Hour"] - df["Start"]).dt.total_seconds() / 60).clip(lower=0)
    df["After Minutes"] = ((df["End"] - df["End Hour"]).dt.total_seconds() / 60).clip(lower=0)

    # Beräkna effekt
    df["Before Power"] = (df["Before Minutes"] / 60) * df["Power"]
    df["After Power"] = (df["After Minutes"] / 60) * df["Power"]
    df["Full Hour Power"] = df["Power"].where(df["Full Hours"] > 0, 0)

    # Ladda arbetsboken och radera fliken om den finns
    try:
        with pd.ExcelWriter(file_path, mode="a", engine="openpyxl") as writer:
            book = load_workbook(file_path)
            if output_sheet in book.sheetnames:
                del book[output_sheet]  # Ta bort gammal flik
                book.save(file_path)
    except FileNotFoundError:
        pass  # Om filen inte finns, skapas den ändå

    # Spara resultatet
    with pd.ExcelWriter(file_path, mode="a", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=output_sheet, index=False)

    print(f"Beräkningar klara! Data sparad i flik: {output_sheet}")

# Exempel på körning
calculate_energy_and_power("energidata.xlsx")